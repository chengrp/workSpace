"""
Excel 测试用例生成器 (使用 openpyxl)
"""

import os
from datetime import datetime
from typing import Dict, List, Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    # 创建 dummy 类以避免 NameError
    class Font:
        def __init__(self, **kwargs): pass
    class Alignment:
        def __init__(self, **kwargs): pass
    class PatternFill:
        def __init__(self, **kwargs): pass
    class Border:
        def __init__(self, **kwargs): pass
    class Side:
        def __init__(self, **kwargs): pass


class ExcelGenerator:
    """Excel 测试用例生成器"""

    def __init__(self, template_dir="../templates"):
        self.template_dir = template_dir

        # 样式定义 (仅在 openpyxl 可用时创建)
        if EXCEL_AVAILABLE:
            self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            self.header_font = Font(bold=True, color="FFFFFF", size=11)
            self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            self.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        else:
            self.header_fill = None
            self.header_font = None
            self.header_alignment = None
            self.border = None

    def generate(self, analysis_result: Dict[str, Any], output_path: str) -> None:
        """
        生成 Excel 测试用例文档

        Args:
            analysis_result: 需求分析结果
            output_path: 输出文件路径
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl 未安装，请运行: pip install openpyxl")

        wb = Workbook()

        # 删除默认 sheet
        wb.remove(wb.active)

        # 创建多个 sheet
        self._create_test_cases_sheet(wb, analysis_result)
        self._create_summary_sheet(wb, analysis_result)
        self._create_questions_sheet(wb, analysis_result)
        self._create_scope_sheet(wb, analysis_result)

        # 保存文件
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)

    def _create_test_cases_sheet(self, wb: Workbook, analysis: Dict) -> None:
        """创建测试用例总表"""
        ws = wb.create_sheet("测试用例总表", 0)

        # 表头
        headers = ["用例ID", "类型", "标题", "模块", "优先级", "前置条件", "测试步骤", "预期结果", "测试数据", "状态", "标签"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.border

        # 设置列宽
        column_widths = [12, 12, 30, 15, 10, 20, 40, 30, 15, 10, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # 数据行
        for row, tc in enumerate(analysis.get("test_cases", []), 2):
            ws.cell(row=row, column=1, value=tc.get("id", ""))
            ws.cell(row=row, column=2, value=tc.get("type", ""))
            ws.cell(row=row, column=3, value=tc.get("title", ""))
            ws.cell(row=row, column=4, value="")
            ws.cell(row=row, column=5, value=tc.get("priority", ""))
            ws.cell(row=row, column=6, value="")
            ws.cell(row=row, column=7, value="\n".join(tc.get("steps", [])))
            ws.cell(row=row, column=8, value=tc.get("expected", ""))
            ws.cell(row=row, column=9, value="")
            ws.cell(row=row, column=10, value="设计")
            ws.cell(row=row, column=11, value=f"#{tc.get('type', '')}")

            # 应用边框
            for col in range(1, 12):
                ws.cell(row=row, column=col).border = self.border
                ws.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)

        # 冻结首行
        ws.freeze_panes = "A2"

    def _create_summary_sheet(self, wb: Workbook, analysis: Dict) -> None:
        """创建需求分析摘要"""
        ws = wb.create_sheet("需求分析摘要", 1)

        row = 1

        # 标题
        ws.merge_cells(f'A{row}:D{row}')
        title_cell = ws.cell(row=row, column=1)
        title_cell.value = "需求分析摘要"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        row += 2

        # 需求信息
        title = analysis.get("summary", {}).get("title", "")
        ws.cell(row=row, column=1, value="需求名称:")
        ws.cell(row=row, column=2, value=title)
        row += 1

        ws.cell(row=row, column=1, value="生成时间:")
        ws.cell(row=row, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        row += 2

        # 核心功能
        ws.cell(row=row, column=1, value="核心功能:")
        ws.cell(row=row, column=2, value="\n".join(analysis.get("summary", {}).get("core_functions", [])))
        row += 2

        # 触发条件
        ws.cell(row=row, column=1, value="触发条件:")
        ws.cell(row=row, column=2, value="\n".join(analysis.get("core_elements", {}).get("trigger_conditions", [])))
        row += 2

        # 计时条件
        timings = analysis.get("core_elements", {}).get("timing_conditions", [])
        if timings:
            ws.cell(row=row, column=1, value="计时条件:")
            timing_text = ", ".join([f"{t['value']}{t['unit']}" for t in timings])
            ws.cell(row=row, column=2, value=timing_text)
            row += 2

        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50

    def _create_questions_sheet(self, wb: Workbook, analysis: Dict) -> None:
        """创建待确认事项"""
        ws = wb.create_sheet("待确认事项", 2)

        # 表头
        headers = ["序号", "待确认事项", "优先级", "建议"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment

        # 列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 40

        # 数据
        questions = analysis.get("questions_to_confirm", [])
        for i, q in enumerate(questions, 2):
            ws.cell(row=i, column=1, value=i-1)
            ws.cell(row=i, column=2, value=q)
            ws.cell(row=i, column=3, value="P1")
            ws.cell(row=i, column=4, value="需与产品/开发确认")

            # 边框
            for col in range(1, 5):
                ws.cell(row=i, column=col).border = self.border

        ws.freeze_panes = "A2"

    def _create_scope_sheet(self, wb: Workbook, analysis: Dict) -> None:
        """创建测试范围评估"""
        ws = wb.create_sheet("测试范围评估", 3)

        # 表头
        headers = ["测试类型", "用例数", "优先级分布", "说明"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment

        # 按类型统计
        test_cases_by_type = {}
        for tc in analysis.get("test_cases", []):
            tc_type = tc.get("type", "其他")
            if tc_type not in test_cases_by_type:
                test_cases_by_type[tc_type] = {"total": 0, "p0": 0, "p1": 0}
            test_cases_by_type[tc_type]["total"] += 1
            if tc.get("priority") == "P0":
                test_cases_by_type[tc_type]["p0"] += 1
            elif tc.get("priority") == "P1":
                test_cases_by_type[tc_type]["p1"] += 1

        # 数据行
        row = 2
        total_cases = 0
        for tc_type, stats in test_cases_by_type.items():
            ws.cell(row=row, column=1, value=tc_type)
            ws.cell(row=row, column=2, value=stats["total"])
            ws.cell(row=row, column=3, value=f"P0:{stats['p0']}, P1:{stats['p1']}")
            ws.cell(row=row, column=4, value="覆盖核心场景")

            # 边框
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self.border

            total_cases += stats["total"]
            row += 1

        # 总计
        ws.cell(row=row, column=1, value="总计")
        ws.cell(row=row, column=2, value=total_cases)
        ws.cell(row=row, column=3, value="-")
        ws.cell(row=row, column=4, value=f"共 {total_cases} 个测试用例")

        # 冻结首行
        ws.freeze_panes = "A2"

        # 列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30


if __name__ == "__main__":
    import sys
    import json

    if not EXCEL_AVAILABLE:
        print("❌ openpyxl 未安装")
        print("请运行: pip install openpyxl")
        sys.exit(1)

    if len(sys.argv) < 2:
        print("Usage: python generate_excel.py <analysis.json> <output.xlsx>")
        sys.exit(1)

    with open(sys.argv[1], 'r', encoding='utf-8') as f:
        analysis = json.load(f)

    generator = ExcelGenerator()
    output = sys.argv[2] if len(sys.argv) > 2 else "output/test_cases.xlsx"
    generator.generate(analysis, output)

    print(f"✅ Excel 已生成: {output}")
